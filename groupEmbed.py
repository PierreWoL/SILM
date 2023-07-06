import os
import pickle
import shutil

import numpy as np
from openpyxl.workbook import Workbook
import pandas as pd
import openpyxl
from Utils import mkdir
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd


# 定义文件夹列表

def ExampleAggregate(datafile_path, target):
    os.getcwd()
    files = [fn for fn in os.listdir(datafile_path)]
    source_folder = os.path.join("D:/CurrentDataset/result/starmie/WDC/", target)  # "TableClassExample"
    mkdir(source_folder)
    csv_files = [fn for fn in os.listdir(os.path.join(datafile_path, files[0]))]
    print(csv_files)
    # 遍历每个文件夹
    for csvf in csv_files:
        destination_folder = os.path.join(source_folder, csvf[:-4])
        mkdir(destination_folder)
        for file in files:
            # 构建源文件的路径和目标文件的路径
            source_file = os.path.join(datafile_path, file, csvf)
            LM = ''
            if "roberta" in file:
                LM = "roberta_"
            if "sbert" in file:
                LM = "sbert_"
            destination_file = os.path.join(destination_folder, LM + file.split("_")[-1] + '.csv')
            # 如果源文件存在，则复制到目标文件夹，并重命名
            if os.path.isfile(source_file):
                # 创建目标文件夹（如果不存在）
                os.makedirs(destination_folder, exist_ok=True)
                # 复制文件并重命名
                shutil.copy(source_file, destination_file)
                print(f"Moved '{source_file}' to '{destination_file}'")


def sum_gen(datapath):
    folders = [fn for fn in os.listdir(datapath) if not fn.endswith(".xlsx")]
    for folder in folders:
        # construct folder
        dest = os.path.join(datapath, folder, "sum.xlsx")
        # check if file exists
        if os.path.exists(dest):
            # delete file
            os.remove(dest)
        workbook = Workbook()
        sheet1 = workbook.active
        sheet1.title = "BIRCH"
        csv_files = [fn for fn in os.listdir(os.path.join(datapath, folder)) if fn.endswith("csv")]
        df_s = pd.DataFrame()
        df2_s = pd.DataFrame()
        # 遍历每个csv文件
        for filename in csv_files:
            csv_path = os.path.join(datapath, folder, filename)
            df = pd.read_csv(csv_path, index_col=0)
            df_s[filename[:-4]] = df.loc[['random score', 'ARI', 'purity'], ['BIRCH']]
            df2_s[filename[:-4]] = df.loc[['random score', 'ARI', 'purity'], ['Agglomerative']]
        df_s = df_s.round(3)
        df2_s = df2_s.round(3)
        print(df_s)
        print(df2_s)
        # write the first dataframe into the first sheet
        rows = list(dataframe_to_rows(df_s, index=True, header=True))
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                sheet1.cell(row=r_idx, column=c_idx, value=value)
        # create the second sheet and name it "Agglomerative"
        sheet2 = workbook.create_sheet(title="Agglomerative")
        rows = list(dataframe_to_rows(df2_s, index=True, header=True))
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                sheet2.cell(row=r_idx, column=c_idx, value=value)
        workbook.save(dest)


def sum_clustering(datapath):
    # def metric_all():
    dfs = {}
    dfs['BIRCH'], dfs['Agglomerative'] = {}, {}
    folders = [fn for fn in os.listdir(datapath) if not fn.endswith(".xlsx")]
    for metric in ['random score', 'ARI', 'purity']:
        dfs['BIRCH'][metric] = pd.DataFrame()
        dfs['Agglomerative'][metric] = pd.DataFrame()
    for folder in folders:
        dest = os.path.join(datapath, folder, "sum.xlsx")
        workbook = openpyxl.load_workbook(dest)
        # 选择要操作的 Sheet
        sheet = workbook['BIRCH']
        sheet1 = workbook['Agglomerative']
        rows_to_delete = []

        for row in sheet.iter_rows():
            if all(cell.value is None for cell in row):
                rows_to_delete.append(row)
        for row in rows_to_delete:
            sheet.delete_rows(row[0].row, 1)
        rows_to_delete = []
        for row in sheet1.iter_rows():
            if all(cell.value is None for cell in row):
                rows_to_delete.append(row)

        workbook.save(dest)

        for clustering_algo in ['BIRCH', 'Agglomerative']:
            data = pd.read_excel(dest, sheet_name=clustering_algo, index_col=0)
            for metric in ['random score', 'ARI', 'purity']:
                selected_row = data.loc[[metric]]
                selected_row.index = [folder.split("_")[0]]  # 更改行索引为 'A'
                dfs[clustering_algo][metric] = pd.concat([dfs[clustering_algo][metric], selected_row])

    print(dfs['BIRCH'])
    print(dfs['Agglomerative'])

    for clustering_al, dataframes_dict in dfs.items():
        dest_f = os.path.join(datapath, clustering_al + ".xlsx")
        writer = pd.ExcelWriter(dest_f, engine='openpyxl')
        # 遍历字典，将每个 DataFrame 保存为对应的 sheet
        for sheet_name, df in dataframes_dict.items():
            if sheet_name == 'random score':
                df.to_excel(writer, sheet_name="Rand Index", index=True)
            else:
                df.to_excel(writer, sheet_name=sheet_name, index=True)
        # 保存 Excel 文件
        writer.close()


"""
dpe = "D:/CurrentDataset/result/starmie/WDC/"
ExampleAggregate(os.path.join(dpe,"example"),"TableClassExample")
ExampleAggregate(os.path.join(dpe,"All"),"TableClass")

dp = "D:/CurrentDataset/result/starmie/WDC/TableClass"
sum_gen(dp)
sum_clustering(dp)
"""

from clustering import clustering_results, data_classes, clusteringColumnResults, clustering_hier_results


def matrixs():
    dataset = "WDC"
    data_path = os.getcwd() + "/datasets/" + dataset + "/Test/"
    ground_truth_table = os.getcwd() + "/datasets/" + dataset + "/groundTruth.csv"
    Gt_clusters,aaa=  data_classes(data_path, ground_truth_table)[0],data_classes(data_path, ground_truth_table)[2]
    ground_truth= pd.read_csv(ground_truth_table)

    matrixs_cols = {}
    matrixs_pairs = {}
    keys_cluster = {}
    Gt_clusters_dict={}
    gt_cluster_dicts = {}
    classes = ["People", "Animal"]
    for superclass in classes:

        Gt_clusters1 = {key: value for key, value in Gt_clusters.items() if value == superclass}

        Gt_clusters_dict[superclass] = {row[1]["fileName"][:-4]:row[1]["Label"] for row in ground_truth.iterrows() if row[1]["fileName"][:-4] in Gt_clusters.keys() and  Gt_clusters[row[1]["fileName"][:-4]]== superclass}
        cluster_list = list(set(list(Gt_clusters_dict[superclass].values())))
        gt_cluster_dicts[superclass] = {i:cluster_list.index(i) for i in cluster_list}

        tables_dict = {key: pd.read_csv(os.path.join(data_path, key + ".csv"), lineterminator='\n') for key in
                       Gt_clusters1.keys()}
        keys = list(tables_dict.keys())
        n = len(keys)
        matrixs_cols[superclass] = [[0] * n for _ in range(n)]
        matrixs_pairs[superclass] = [[0] * n for _ in range(n)]
        keys_cluster[superclass] = keys
        for i in range(n):
            for j in range(n):
                key1 = keys[i]
                key2 = keys[j]
                # print(key1,key2,len(tables_dict[key1].columns),len(tables_dict[key2].columns) )
                value_i = len(tables_dict[key1].columns) +len(tables_dict[key2].columns)
                matrixs_cols[superclass][i][j] = value_i
                matrixs_pairs[superclass][i][j] = (key1, key2)
    return  Gt_clusters_dict,gt_cluster_dicts,keys_cluster,matrixs_cols,matrixs_pairs

def calculate_matrix(file):
    global j
    Gt_clusterss, Gt_cluster_dicts ,keys_cluster,matrixs_cols,matrixs_pairs = matrixs()
    clusters = [fn for fn in os.listdir(file)]
    for cluster_file in clusters:
        F = open(os.path.join(file, cluster_file), 'rb')
        clusters_result = pickle.load(F)
        # Example: {'People': {'BIRCH': {38: ['SOTAB_104.0', 'SOTAB_125.0'
        for cluster_name in clusters_result.keys():
            # clustering_algo
            if cluster_name !="People":
                continue
            Gt_clusters = Gt_clusterss[cluster_name]
            Gt_cluster_dict = Gt_cluster_dicts[cluster_name]

            matrix = matrixs_cols[cluster_name]
            matrixs_pair = matrixs_pairs[cluster_name]
            keys  = keys_cluster[cluster_name]

            cluster = clusters_result[cluster_name]
            methods_metrics = {}
            dictss = {}
            for clustering_algo in cluster.keys():
                """if clustering_algo == "Agglomerative":
                    continue"""

                dicts = {}
                result = cluster[clustering_algo]
                matched_cols = {}
                for column_cluster, columns in result.items():
                    tables = list(set([col.split(".")[0] for col in columns]))
                    cols = set([col for col in columns])
                    #print(column_cluster, tables)
                    combinations = [(key1, key2) for key1 in tables for key2 in tables if key1 != key2]
                    #combinations = [(tables[i], tables[j]) for i in range(0,len(tables)) for j in range(j,len(tables))]
                    #print("combinations",combinations)
                    for pair in combinations:
                        N_matchedColumns = len([col for col in cols if col.split(".")[0] in pair])
                        if pair in matched_cols.keys():
                            matched_cols[pair] += N_matchedColumns
                        else:
                            matched_cols[pair] = N_matchedColumns

                    """for pair in combinations:
                        if pair in matched_cols.keys():
                            matched_cols[pair] += 1
                        else:
                            matched_cols[pair] = 1"""
                empty_matrix = [[0 for _ in row] for row in matrix]
                for row_idx, row in enumerate(matrixs_pair):
                    for col_idx, value in enumerate(row):
                        if value in matched_cols.keys():
                            empty_matrix[row_idx][col_idx] = matched_cols[value]
                #print("matches", empty_matrix)
                #print("distinc", matrix)
                for row_idx, row in enumerate(matrixs_pair):
                    for col_idx, value in enumerate(row):
                        if row_idx!=col_idx:
                            if matrix[row_idx][col_idx]==empty_matrix[row_idx][col_idx]:
                                empty_matrix[row_idx][col_idx] = 0
                            else:
                                empty_matrix[row_idx][col_idx] =  1- (empty_matrix[row_idx][col_idx]/
                                                                  (matrix[row_idx][col_idx]-empty_matrix[row_idx][col_idx]))
                        else:
                            empty_matrix[row_idx][col_idx] = 0
                print("empty_matrix", empty_matrix)

                dict_file = {}
                method_metrics = {}
                metric_value_df =pd.DataFrame(columns=["MI", "NMI", "AMI", "random score", "ARI", "FMI", "purity"])
                for algo in ["GMM","KMeans",'Agglomerative']:#"BIRCH",
                    print(algo)
                    try:
                        for i in range(0, 3):
                            empty_matrix = np.array(empty_matrix)
                            cluster_dict, metric_dict = clustering_hier_results(empty_matrix, keys, Gt_clusters,
                                                                            Gt_cluster_dict, algo)
                            #print("cluster_dict", cluster_dict)
                            metric_df = pd.DataFrame([metric_dict])
                            metric_value_df = pd.concat([metric_value_df, metric_df])
                            dict_file[algo + "_" + str(i)] = cluster_dict

                        mean_metric = metric_value_df.mean()
                        method_metrics[algo] = mean_metric
                    except (ValueError,AttributeError) as e:
                        print(e)
                        continue

                e_df = pd.DataFrame()
                for i, v in method_metrics.items():
                    e_df = pd.concat([e_df, v.rename(i)], axis=1)
                store_path = os.getcwd() + "/result/" + "starmie" + "/" + "WDC" + "/" + "subclass/" + cluster_file[
                                                                                                          :-20] + "/" + cluster_name
                mkdir(store_path)
                e_df.to_csv(os.path.join(store_path, clustering_algo + '_metrics.csv'), encoding='utf-8')
                dicts[file] = dict_file
                dictss[clustering_algo] = dicts
                methods_metrics[clustering_algo] = method_metrics
            datafile_path = os.getcwd() + "/result/" + "starmie"+ "/" + "WDC"+ "/" +"subclass/examples/"+cluster_file+"/"+cluster_name+"/"
            mkdir(datafile_path)
            with open(datafile_path + 'cluster_dict.pickle', 'wb') as handle:
                pickle.dump(dictss, handle, protocol=pickle.HIGHEST_PROTOCOL)




file = "D:/CurrentDataset/result/embedding/starmie/vectors/WDC/cluster"
calculate_matrix(file)


# data_path = "D:/CurrentDataset/datasets/WDC/Test"



