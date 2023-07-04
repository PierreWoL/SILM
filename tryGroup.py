import pandas as pd
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

groundTruth_file = os.getcwd() + "/datasets/WDC/column_gt.xlsx"
ground_truth_df = pd.read_excel(groundTruth_file, sheet_name=0)
unique_tables = ground_truth_df['Table_Cluster_Label'].unique()
print(unique_tables)
column_cluster = {}
"""gt_clusters: tablename.Columnname:corresponding label dictionary. e.g.{'SOTAB_0.a1': 'Game', 'SOTAB_0.b1': 'date',...}
    ground_t: label and its tables. e.g. {'Game': ['SOTAB_0.a1'], 'Newspaper': ['SOTAB_0.b1', 'SOTAB_0.c1'],...}
    gt_cluster_dict: dictionary of index: label
    like {'Game': 0, 'date': 1, ...}
"""
for table in unique_tables:
    grouped_df = ground_truth_df[ground_truth_df['Table_Cluster_Label'] == table]
    unique_column_cluster = grouped_df['Column_Cluster_Label'].unique()

    column_cluster[table] = {}

    for cluster in unique_column_cluster:
        # column_cluster[table][cluster] = []
        Cluster_col = []
        filtered_rows = grouped_df[grouped_df['Column_Cluster_Label'] == cluster]
        for index, row in filtered_rows.iterrows():
            Cluster_col.append(f"{row['Source_Dataset']}.{row['column1']}")
            Cluster_col.append(f"{row['Target_Dataset']}.{row['column2']}")
        column_cluster[table][cluster] = list(set(Cluster_col))
print(len(column_cluster))
# Print the resulting dictionary
print(column_cluster)
# Create a workbook and select the active sheet
workbook = Workbook()
sheet = workbook.active
# sheet = workbook.create_sheet(title="sheet2")
new_df_dict = {}
df_saveA = pd.DataFrame([(key, len(values)) for key, values in column_cluster.items()],
                        columns=['tableCategory', 'columnsCategoryNumber'])
df_saveA.to_csv(os.path.join(os.getcwd(), "datasets/WDC/aggre.csv"))
for table, clusters in column_cluster.items():
    df_save = pd.DataFrame([(key, len(values)) for key, values in clusters.items()],
                           columns=['columnCategory', 'columnsNumber'])
    # new_df_dict[table] = df_save
    # Write the table name as a merged cell
    sheet.merge_cells(start_row=sheet.max_row + 2, start_column=1, end_row=sheet.max_row + 2, end_column=2)
    sheet.cell(row=sheet.max_row, column=1, value=table)

    # Write the table data to the sheet
    for row in dataframe_to_rows(df_save, index=False, header=True):
        sheet.append(row)

# Save the workbook as an Excel file
#workbook.save(os.path.join(os.getcwd(), "datasets/WDC/outputAggre.xlsx"))

for table, clusters in column_cluster.items():
    print(f"Table Category: {table}")
    ground_t = clusters
    gt_clusters = {}
    ck = list(clusters.keys())
    gt_cluster_dict = {ck[i]: i for i in range(0, len(ck))}
    for cluster, values in clusters.items():
        print(f" {cluster} {len(values)}")
